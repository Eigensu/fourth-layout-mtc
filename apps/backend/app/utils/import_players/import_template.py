"""Template generation utilities for player import"""
import io
import csv
from typing import BinaryIO, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation


# Define standard columns
TEMPLATE_COLUMNS = [
    "name",
    "team", 
    "points",
    "slot_code",
    "slot_name",
    "gender",
    "mobile",
    "status",
    "image_url",
    "matches",
    "runs",
    "wickets",
]

TEMPLATE_HEADERS = [
    "Name",
    "Team",
    "Points",
    "Slot Code",
    "Slot Name",
    "Gender",
    "Mobile",
    "Status",
    "Image URL",
    "Matches",
    "Runs",
    "Wickets",
]


async def generate_xlsx_template(slot_codes: Optional[list[str]] = None) -> BinaryIO:
    """
    Generate XLSX template with data validations
    
    Args:
        slot_codes: Optional list of slot codes for dropdown
        
    Returns:
        Binary file-like object with XLSX content
    """
    wb = Workbook()
    ws = wb.active
    
    if ws is None:
        raise ValueError("Failed to create worksheet")
    
    ws.title = "Players"
    
    # Style for header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_idx, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header  # type: ignore[assignment]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Set column widths
    ws.column_dimensions["A"].width = 20  # Name
    ws.column_dimensions["B"].width = 25  # Team
    ws.column_dimensions["C"].width = 10  # Points
    ws.column_dimensions["D"].width = 15  # Slot Code
    ws.column_dimensions["E"].width = 15  # Slot Name
    ws.column_dimensions["F"].width = 12  # Gender
    ws.column_dimensions["G"].width = 16  # Mobile
    ws.column_dimensions["H"].width = 12  # Status
    ws.column_dimensions["I"].width = 30  # Image URL
    ws.column_dimensions["J"].width = 12  # Matches
    ws.column_dimensions["K"].width = 12  # Runs
    ws.column_dimensions["L"].width = 12  # Wickets
    
    # Data validation for Gender (column F)
    gender_dv = DataValidation(
        type="list",
        formula1='"male,female"',
        allow_blank=False
    )
    gender_dv.error = "Please select 'male' or 'female'"
    gender_dv.errorTitle = "Invalid Gender"
    ws.add_data_validation(gender_dv)
    gender_dv.add(f"F2:F5000")
    
    # Data validation for Status (column H)
    status_dv = DataValidation(
        type="list",
        formula1='"Active,Inactive,Injured"',
        allow_blank=True
    )
    status_dv.error = "Please select a valid status"
    status_dv.errorTitle = "Invalid Status"
    ws.add_data_validation(status_dv)
    status_dv.add(f"H2:H5000")
    
    # Data validation for Slot Code (column D) - MEN or WOMEN only
    slot_dv = DataValidation(
        type="list",
        formula1='"MEN,WOMEN"',
        allow_blank=False
    )
    slot_dv.error = "Please select 'MEN' or 'WOMEN'"
    slot_dv.errorTitle = "Invalid Slot"
    ws.add_data_validation(slot_dv)
    slot_dv.add(f"D2:D5000")
    
    # Add example rows for both genders
    example_rows = [
        # Men's player example
        [
            "Virat Kohli",
            "Mumbai Indians Men",
            850,
            "MEN",
            "Men",
            "male",
            "9876543210",
            "Active",
            "https://example.com/virat.jpg",
            120,
            5420,
            0,
        ],
        # Women's player example
        [
            "Smriti Mandhana",
            "Mumbai Indians Women",
            720,
            "WOMEN",
            "Women",
            "female",
            "9876543211",
            "Active",
            "https://example.com/smriti.jpg",
            85,
            3240,
            0,
        ],
    ]
    
    for row_idx, example_row in enumerate(example_rows, start=2):
        for col_idx, value in enumerate(example_row, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value
    
    # Add instructions in a new sheet
    instructions = wb.create_sheet("Instructions")
    instructions.column_dimensions["A"].width = 80
    
    instruction_text = [
        ("Player Import Template Instructions", True),
        ("", False),
        ("IMPORTANT - New Team Selection System:", True),
        ("• Only 2 slots available: MEN and WOMEN", False),
        ("• Team names MUST include gender (e.g., 'Mumbai Indians Men', 'Mumbai Indians Women')", False),
        ("• Each team should have exactly 5 players", False),
        ("• Gender field is REQUIRED for all players", False),
        ("", False),
        ("Required Fields:", True),
        ("• Name: Player's full name (required)", False),
        ("• Team: Team name with gender suffix (required, e.g., 'Team A Men')", False),
        ("• Points: Player points (required, numeric)", False),
        ("• Slot Code: Must be 'MEN' or 'WOMEN' (required)", False),
        ("• Gender: Must be 'male' or 'female' (required)", False),
        ("", False),
        ("Optional Fields:", True),
        ("• Slot Name: Auto-filled based on Slot Code", False),
        ("• Mobile: Contact number for the player", False),
        ("• Status: Active, Inactive, or Injured (default: Active)", False),
        ("• Image URL: Player image URL", False),
        ("• Stats: Matches, Runs, Wickets (stored as player stats)", False),
        ("", False),
        ("Team Selection Rules:", True),
        ("• Total squad: 16 players (12 men + 4 women)", False),
        ("• Maximum 3 players from any single team (per gender)", False),
        ("• Example: Can select 3 from 'Mumbai Indians Men' AND 3 from 'Mumbai Indians Women'", False),
        ("", False),
        ("Tips:", True),
        ("• Use the dropdown menu for Gender, Status, and Slot Code", False),
        ("• Ensure team names are consistent (same spelling, capitalization)", False),
        ("• Delete the example rows before importing", False),
        ("• Save file as .xlsx format", False),
        ("• Maximum 5,000 rows per file", False),
    ]
    
    for row_idx, (text, bold) in enumerate(instruction_text, start=1):
        cell = instructions.cell(row=row_idx, column=1)
        cell.value = text
        if bold:
            cell.font = Font(bold=True, size=12)
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_csv_template() -> str:
    """Generate CSV template as string"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(TEMPLATE_COLUMNS)
    
    # Write example rows
    writer.writerow([
        "Virat Kohli",
        "Mumbai Indians Men",
        850,
        "MEN",
        "Men",
        "male",
        "9876543210",
        "Active",
        "https://example.com/virat.jpg",
        120,
        5420,
        0,
    ])
    writer.writerow([
        "Smriti Mandhana",
        "Mumbai Indians Women",
        720,
        "WOMEN",
        "Women",
        "female",
        "9876543211",
        "Active",
        "https://example.com/smriti.jpg",
        85,
        3240,
        0,
    ])
    
    return output.getvalue()
